"""
Export handler.
Generates an advanced Excel file with all user transactions and debts.
"""

import io
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from aiogram import Router, F
from aiogram.filters import Command
from aiogram.types import Message, BufferedInputFile, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from src.database import transactions_collection, get_active_debts, get_user
from src.states import ExportStates

router = Router()

def get_export_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📅 Bu oy", callback_data="export_this_month"),
            InlineKeyboardButton(text="🗓 O'tgan oy", callback_data="export_last_month")
        ],
        [
            InlineKeyboardButton(text="📊 Oxirgi 3 oy", callback_data="export_last_3_months"),
            InlineKeyboardButton(text="♾ Barcha vaqt", callback_data="export_all_time")
        ],
        [
            InlineKeyboardButton(text="⚙️ Boshqa sana (Custom)", callback_data="export_custom")
        ]
    ])

@router.message(Command("excel"))
@router.message(F.text == "📥 Yuklab olish")
async def show_export_menu(message: Message, state: FSMContext):
    await state.clear()
    await message.answer(
        "📥 <b>Hisobot yuklab olish</b>\n\nQaysi davr uchun hisobot kerakligini tanlang:",
        reply_markup=get_export_keyboard(),
        parse_mode="HTML"
    )

@router.callback_query(F.data == "export_custom")
async def ask_custom_date(callback: CallbackQuery, state: FSMContext):
    await state.set_state(ExportStates.waiting_for_custom_date)
    await callback.message.edit_text(
        "⚙️ <b>Maxsus sana</b>\n\n"
        "Qaysi sanalar oralig'idagi hisobot kerak?\n"
        "Iltimos, sanani quyidagi formatda kiriting:\n\n"
        "<code>DD.MM.YYYY - DD.MM.YYYY</code>\n"
        "Masalan: <code>01.10.2023 - 31.10.2023</code>",
        parse_mode="HTML"
    )

@router.message(ExportStates.waiting_for_custom_date)
async def process_custom_date(message: Message, state: FSMContext):
    text = message.text.strip()
    try:
        parts = text.split("-")
        if len(parts) != 2:
            raise ValueError
            
        start_str = parts[0].strip()
        end_str = parts[1].strip()
        
        start_date = datetime.strptime(start_str, "%d.%m.%Y")
        end_date = datetime.strptime(end_str, "%d.%m.%Y").replace(hour=23, minute=59, second=59)
        
        if start_date > end_date:
            await message.answer("❌ Boshlanish sanasi tugash sanasidan kichik bo'lishi kerak. Qayta kiriting:")
            return
            
        await state.clear()
        period_text = f"{start_str} - {end_str}"
        await generate_and_send_excel(message, message.from_user.id, start_date, end_date, period_text)
        
    except Exception:
        await message.answer(
            "❌ <b>Xato format!</b>\nIltimos, to'g'ri formatda kiriting:\n"
            "<code>01.10.2023 - 31.10.2023</code>",
            parse_mode="HTML"
        )

@router.callback_query(F.data.startswith("export_"))
async def process_export_period(callback: CallbackQuery):
    action = callback.data
    if action == "export_custom":
        return # Handled above
        
    now = datetime.now()
    start_date = None
    end_date = now
    period_text = "Barcha vaqt"
    
    if action == "export_this_month":
        start_date = datetime(now.year, now.month, 1)
        period_text = now.strftime("%m.%Y")
    elif action == "export_last_month":
        first_day_this_month = datetime(now.year, now.month, 1)
        last_day_last_month = first_day_this_month - timedelta(days=1)
        start_date = datetime(last_day_last_month.year, last_day_last_month.month, 1)
        end_date = last_day_last_month.replace(hour=23, minute=59, second=59)
        period_text = start_date.strftime("%m.%Y")
    elif action == "export_last_3_months":
        start_date = now - timedelta(days=90)
        period_text = "Oxirgi 3 oy"
        
    await callback.message.delete()
    await generate_and_send_excel(callback.message, callback.from_user.id, start_date, end_date, period_text)


async def generate_and_send_excel(message: Message, user_id: int, start_date, end_date, period_text: str):
    wait_msg = await message.answer("⏳ <i>Hisobot tayyorlanmoqda, kuting...</i>", parse_mode="HTML")
    
    # Query logic
    query = {"telegram_id": user_id}
    if start_date and end_date:
        query["created_at"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["created_at"] = {"$gte": start_date}
        
    cursor = transactions_collection.find(query).sort("created_at", 1) # ASC for balance tracking
    transactions = await cursor.to_list(length=100000)
    
    user = await get_user(user_id)
    user_name = user.get("full_name", "User").split()[0] # First name
    
    # Calculate summary
    total_income = 0
    total_expense = 0
    category_sums = {}
    
    # Format transactions for display
    formatted_txs = []
    current_balance = 0
    
    for tx in transactions:
        t_type = tx.get("type", "").lower()
        amount = tx.get("amount", 0)
        cat = tx.get("category", "")
        
        if t_type == "kirim":
            total_income += amount
            current_balance += amount
        elif t_type == "chiqim":
            total_expense += amount
            current_balance -= amount
            # Track categories for expenses
            category_sums[cat] = category_sums.get(cat, 0) + amount
        elif t_type == "qarz":
            pass # Usually handled separate or doesn't affect standard income/expense breakdown
            
        formatted_txs.append({
            "date": tx.get("date", ""),
            "type": t_type.capitalize(),
            "amount": amount,
            "currency": tx.get("currency", "UZS"),
            "category": cat,
            "desc": tx.get("description", ""),
            "balance": current_balance
        })
        
    net_balance = total_income - total_expense
    
    top_category = "Yo'q"
    if category_sums:
        top_category = max(category_sums.items(), key=lambda x: x[1])[0]

    # Initialize Excel Workbook
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Light green
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light red
    yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Light yellow
    
    thin_border = Border(left=Side(style='thin', color='E5E7EB'), 
                         right=Side(style='thin', color='E5E7EB'), 
                         top=Side(style='thin', color='E5E7EB'), 
                         bottom=Side(style='thin', color='E5E7EB'))

    # ─── SHEET 1: Umumiy xulosa ───
    ws_summary = wb.active
    ws_summary.title = "Umumiy xulosa"
    
    ws_summary.merge_cells('A1:D1')
    title_cell = ws_summary['A1']
    title_cell.value = f"MOLIYAVIY XULOSA ({period_text})"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = header_fill
    title_cell.alignment = center_align
    
    metrics = [
        ("Jami Kirim:", total_income),
        ("Jami Chiqim:", total_expense),
        ("Sof Qoldiq:", net_balance),
        ("Eng ko'p xarajat:", top_category)
    ]
    
    for row, (label, val) in enumerate(metrics, 3):
        ws_summary.cell(row=row, column=1).value = label
        ws_summary.cell(row=row, column=1).font = Font(bold=True)
        
        c2 = ws_summary.cell(row=row, column=2)
        c2.value = val
        if label == "Jami Kirim:": c2.font = Font(color="16A34A", bold=True)
        elif label == "Jami Chiqim:": c2.font = Font(color="DC2626", bold=True)
        elif label == "Sof Qoldiq:": c2.font = Font(color="16A34A" if val >= 0 else "DC2626", bold=True)

    ws_summary.cell(row=8, column=1).value = "Kategoriyalar bo'yicha chiqim:"
    ws_summary.cell(row=8, column=1).font = Font(bold=True)
    
    row_idx = 9
    for cat, summ in sorted(category_sums.items(), key=lambda x: x[1], reverse=True):
        ws_summary.cell(row=row_idx, column=1).value = cat
        ws_summary.cell(row=row_idx, column=2).value = summ
        row_idx += 1

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20

    # ─── SHEET 2: Barcha tranzaksiyalar ───
    ws_tx = wb.create_sheet(title="Barcha tranzaksiyalar")
    
    headers_tx = ["№", "Sana", "Tur", "Summa", "Valyuta", "Kategoriya", "Izoh", "Balans (keyin)"]
    for col, h in enumerate(headers_tx, 1):
        cell = ws_tx.cell(row=1, column=col)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    # Reverse to show newest first? User requirement: chronological makes sense for running balance, but newest first is better for reading. 
    # Let's put newest first, but the running balance is correct.
    formatted_txs.reverse()
    
    for row, tx in enumerate(formatted_txs, 2):
        ws_tx.cell(row=row, column=1).value = len(formatted_txs) - row + 2
        ws_tx.cell(row=row, column=2).value = tx["date"]
        ws_tx.cell(row=row, column=3).value = tx["type"]
        ws_tx.cell(row=row, column=4).value = tx["amount"]
        ws_tx.cell(row=row, column=5).value = tx["currency"]
        ws_tx.cell(row=row, column=6).value = tx["category"]
        ws_tx.cell(row=row, column=7).value = tx["desc"]
        ws_tx.cell(row=row, column=8).value = tx["balance"]
        
        # Apply colors and borders
        fill_color = None
        if tx["type"] == "Kirim": fill_color = green_fill
        elif tx["type"] == "Chiqim": fill_color = red_fill
        elif tx["type"] == "Qarz": fill_color = yellow_fill
        
        for col in range(1, 9):
            c = ws_tx.cell(row=row, column=col)
            c.border = thin_border
            if fill_color:
                c.fill = fill_color

    ws_tx.column_dimensions['A'].width = 5
    ws_tx.column_dimensions['B'].width = 15
    ws_tx.column_dimensions['C'].width = 10
    ws_tx.column_dimensions['D'].width = 15
    ws_tx.column_dimensions['E'].width = 10
    ws_tx.column_dimensions['F'].width = 20
    ws_tx.column_dimensions['G'].width = 30
    ws_tx.column_dimensions['H'].width = 15

    # ─── SHEET 3: Qarzlar ───
    ws_debts = wb.create_sheet(title="Qarzlar")
    
    ws_debts.merge_cells('A1:E1')
    d_title1 = ws_debts['A1']
    d_title1.value = "BERISHIM KERAK BO'LGAN QARZLAR"
    d_title1.font = Font(bold=True, color="FFFFFF")
    d_title1.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    d_title1.alignment = center_align
    
    headers_d = ["Ism", "Qolgan summa", "Valyuta", "Sana", "Muddat", "Holat"]
    for col, h in enumerate(headers_d, 1):
        cell = ws_debts.cell(row=2, column=col)
        cell.value = h
        cell.font = Font(bold=True)
    
    berishim_kerak = await get_active_debts(user_id, "olgan") # Olingan qarzni berish kerak
    
    current_row = 3
    for d in berishim_kerak:
        ws_debts.cell(row=current_row, column=1).value = d.get("person", "")
        ws_debts.cell(row=current_row, column=2).value = d["amount"] - d.get("paid_amount", 0)
        ws_debts.cell(row=current_row, column=3).value = d.get("currency", "UZS")
        ws_debts.cell(row=current_row, column=4).value = d.get("date", "")
        ws_debts.cell(row=current_row, column=5).value = d.get("due_date", "Nomalum")
        ws_debts.cell(row=current_row, column=6).value = "Aktiv"
        current_row += 1
        
    current_row += 2
    ws_debts.merge_cells(f'A{current_row}:E{current_row}')
    d_title2 = ws_debts[f'A{current_row}']
    d_title2.value = "OLISHIM KERAK BO'LGAN QARZLAR"
    d_title2.font = Font(bold=True, color="FFFFFF")
    d_title2.fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    d_title2.alignment = center_align
    current_row += 1
    
    for col, h in enumerate(headers_d, 1):
        cell = ws_debts.cell(row=current_row, column=col)
        cell.value = h
        cell.font = Font(bold=True)
    current_row += 1
    
    olishim_kerak = await get_active_debts(user_id, "bergan") # Berilgan qarzni olish kerak
    for d in olishim_kerak:
        ws_debts.cell(row=current_row, column=1).value = d.get("person", "")
        ws_debts.cell(row=current_row, column=2).value = d["amount"] - d.get("paid_amount", 0)
        ws_debts.cell(row=current_row, column=3).value = d.get("currency", "UZS")
        ws_debts.cell(row=current_row, column=4).value = d.get("date", "")
        ws_debts.cell(row=current_row, column=5).value = d.get("due_date", "Nomalum")
        ws_debts.cell(row=current_row, column=6).value = "Aktiv"
        current_row += 1

    ws_debts.column_dimensions['A'].width = 20
    ws_debts.column_dimensions['B'].width = 15
    ws_debts.column_dimensions['C'].width = 10
    ws_debts.column_dimensions['D'].width = 15
    ws_debts.column_dimensions['E'].width = 15
    ws_debts.column_dimensions['F'].width = 10

    # ─── Save and Send ───
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    file_period = period_text.replace(" ", "").replace("/", "-")
    file_name = f"Somly_{user_name}_{file_period}.xlsx"
    file = BufferedInputFile(stream.read(), filename=file_name)
    
    await wait_msg.delete()
    await message.answer_document(
        document=file,
        caption="📊 <b>Hisobotingiz tayyor!</b>",
        parse_mode="HTML"
    )
